import openpyxl
import classes as cl
import variable as v
from db.db import ConnectBD
from collections import Counter

client = ConnectBD(
    user_name="superadmin",
    password="02Transfer_fer20",
    host="79.174.88.142",
    port="19490",
    database_name="BDTransferBot"
)

# client.__drop_table__()

def find_person(person_id: int) -> cl.User:

    """
    Данная функция возвращает экземпляр класса User с нужным пользователем.
    Идентификация происходит по его id.
    """

    if client.find_user(person_id):
        return cl.User(
            dict(
                zip(
                    v.personal_data_fields,
                    client.find_user(person_id)[0]
                )
            )
        )
    
    else:
        return None

    
def add_application(data_user: dict) -> None:

    """
    Данная функция принимает словарь с данными о клиенте и добавляет её в базу данных.
    """

    client.add_user(
        user_id= data_user["id"],
        name= data_user["name"],
        surname= data_user["surname"],
        patronymic= data_user["patronymic"],
        phone_number= data_user["phone_number"]
    )


def enrollment_cash(person_id: int, price: int, operator: str) -> None:
    """
    Данная функция зачисления/снятия определённой суммы price на виртуальный счёт пользвателя, который определяется по user_id.
    """

    client.enrollment_cash(user_id=person_id, price=price, operator=operator)


def get_full_info_personal() -> list[tuple]:

    """
    Данная функция возвразает список списков с персональными данными.
    """

    return client.get_all_info_personal()


def created_xlsx_persons_data() -> bytes:

    """
    Данная функция создаёт Excel файл и информацией о персональных данных
    и возвращает его в битовом варианте.
    """

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    sheet = workbook.create_sheet(title="Персональные данные")

    sheet.append(v.personal_data_fields)
    for elements in client.get_all_info_personal():
        sheet.append(elements)

    workbook.save(filename="db/persons_data.xlsx")

    with open(file="db/persons_data.xlsx", mode="rb") as file:
        bytes_xlsx = file.read()

    return bytes_xlsx


def dowload_info_xlsx(file_bytes: bytes) -> None:

    """
    Данная функция скачивает файл из Телеграмма. 
    """
    
    with open(file="db/dowload_info_personal.xlsx", mode="wb") as file:
        file.write(file_bytes)

    send_db()


def send_db() -> None:

    """
    Данная функция читает информацию из Excel файла и возвращает список списков.
    Отправляет информацию в СУБД.
    """
    
    workbook = openpyxl.open(filename="db/dowload_info_personal.xlsx", read_only=True)["Персональные данные"]
    
    """
    Каждое значение таблицы Excel относится к классу.
    Конструкция ниже используется для того чтобы
    перевести экземпляр класса в читаемый вид
    
    """
    data = list(map(lambda string: list(map(lambda value: value.value, string)), workbook))[1:]

    for row in data:
        
        try:
            row_2 = {
                    "id": int(row[0]),
                    "Статус": v.category_id[row.pop(4)],
                    "id_data": int(row[0]),
                    "Баланс": row.pop(4),
                }
        except KeyError:
            return
        
        client.update_personal_data(li=row)
        client.update_personal(li=list(row_2.values()))


def send_info_orders(data: dict) -> None:

    """
    Данная функция отправлять информация в СУБД для дальнейшей обработки.
    Переводит словари в списки.
    """

    client.add_orders(li=list(data.values()))


def get_info_orders() -> list[tuple]:

    """
    Данная функция получает информацию о заказах из базы данных.
    """

    iterable = client.show_info_orders()
    
    update_list = []
    list_index_orders = map(lambda x: x[1], iterable)

    for order_id, count_repeat in Counter(list_index_orders).items():
        for info_order in iterable:
            if order_id == info_order[1]:
                temperary_list = list(info_order)
                temperary_list.append(count_repeat)
                update_list.append(temperary_list)
                break

    for element in update_list:
        if element[0] == None:
            element[-1] -= 1

    return update_list


def create_order_fields() -> bytes:

    """
    Данная функция создаёт Excel файл с пунктами о заказе,
    и возвращает его в битовом варианте
    """

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet(title="Данные о заказе")
    sheet = workbook["Данные о заказе"]

    for elements in v.order_data_fields:
        sheet.append([elements])

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 35  

    workbook.save(filename="db/order_fields.xlsx")

    with open(file="db/order_fields.xlsx", mode="rb") as file:
        bytes_xlsx = file.read()

    return bytes_xlsx


def reader_order_fields(file_bytes: bytes) -> list:

    """
    Данная функция преобразует битовый вариант файла и изымает информацию
    о заказе. Возвращает её в виде списка.
    """

    with open(file="db/order_fields.xlsx", mode="wb") as file:
        file.write(file_bytes)

    workbook = openpyxl.open(filename="db/order_fields.xlsx", read_only=True)
    sheet = workbook["Данные о заказе"]
    
    data_list =[]

    for element in sheet:
        if bool(element[1].value):
            data_list.append(str(element[1].value))

    return data_list


def find_info_order(order_id: int) -> cl.Order:
    
    """
    Данная функция ищет заказ по его id в базе данных и возвращает его
    в качестве экземпляра класса Order.
    """

    active_load_man = "None"
    for order in get_info_orders():
        if int(order_id) == int(order[1]):
            active_load_man = order[-1]
            
    return cl.Order(li=client.find_order(order_id=order_id)[0], active_loader_man=active_load_man)


def add_order_persons(order_id: int, user_id):
    
    """
    Данная функция отправляет данные в СУБД по принятому заказу.
    """
        # Добавления грузчика на заказ
    client.add_orders_personals(order_id=order_id, user_id=user_id)

        # Проверка на укомплектованность заказа
    order = find_info_order(order_id=order_id)
    if order.max_count_loader_man == order.active_loader_man:
        client.update_status_order(order_id=order_id)


def get_order_personal_info() -> list[tuple]:

    """
    Данная функция получает и возвращает данные об активных заказов из базы данных.
    """

    return client.get_order_personal_info()


def delete_active_orders(user_id: int, order_id: int) -> None:

    """
    Данная функция отправляет данные в СУБД по удалению активного заказа.
    """
        # Удаление грузчика из списка активных.
    client.delete_active_order(user_id=user_id)

        # Проверка на антиукомплектование заказа для дальнейшего удаления.
    order = find_info_order(order_id=order_id)
    if order.active_loader_man == 0 and order.status == "Принят":
        client.delete_order(order_id=order_id)


def delete_order_admin(order_id: int) -> None:
    
    """
    Данная функция удаляет заказы. Идентификация происходит по его id.
    """

    client.delete_order(order_id=order_id)


def active_load_man(order_id: int) -> tuple[int, int]:
    
    """
    Данная функция возвращает информацию о степени укомплектованности заказа.
    """

    for basa_order in get_info_orders():
        if order_id == basa_order[1]:
            load_man = basa_order[-6]
            active_load_man = basa_order[-1]
            return active_load_man, load_man
    